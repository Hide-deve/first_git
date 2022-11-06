
import openpyxl

#excelブックを読み込む
wb = openpyxl.load_workbook("出社在宅集計表_人事部.xlsx")
print(wb)

#シートを代入
ws = wb["4月"]
print(ws)

#セルを代入 rowとcolumnを指定する .valueでセルの値を代入
cell = ws.cell(row=2 ,column= 9).value
print(cell)

#5行目5列目に9999を書き込み、表示
ws.cell(row=5 ,column= 5).value = 9999
cell02 = ws.cell(row=5 ,column= 5).value
print(cell02)

#excelを()の名前で保存
wb.save("人事在宅出勤02.xlsx")

#列番号を取得
openpyxl.utils.column_index_from_string("AE")

#列番号からExcelの列文字を取得する
openpyxl.utils.get_column_letter(100)

#最終列の列番号を取得する ワークシートのメソッドを利用する 引数はなし
ws.max_column

#新しいシートを()の名前で作成する
ws = wb.create_sheet(title="新規作成シート")

